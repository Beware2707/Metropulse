"""Load curated station accessibility + parking facilities from a Delhi
Transport Stack xlsx export and match each row to a loaded GTFS stop.

The xlsx (``dmrc_station_details_with_parking.xlsx``) carries its own
``station_code``, but the GTFS static feed already loaded into this
database (see :class:`~metropulse.infrastructure.db.models.Stop`) has an
empty ``stop_code`` column, so the two datasets cannot be joined on code.
Instead, stations are matched by normalized name, falling back to nearest
GTFS stop by coordinate for rows whose name doesn't line up (renamed/typo'd
stations, mostly). Every unmatched row is reported back to the caller --
never silently dropped -- via :attr:`MatchResult.unmatched`.

Pure logic + I/O only: no FastAPI/CLI imports here, so the matching
algorithm stays unit-testable in isolation.
"""

from __future__ import annotations

import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook

from metropulse.application.commuter.geo_matching import haversine_meters, normalize_station_name
from metropulse.infrastructure.db.base import SessionFactory
from metropulse.infrastructure.db.commuter_models import StationFacility
from metropulse.infrastructure.db.commuter_repositories import StationFacilityRepository
from metropulse.infrastructure.db.models import Stop
from metropulse.infrastructure.db.repositories import StopRepository

_COORDINATE_MATCH_RADIUS_M = 500.0

# Numeric capacity fields: the source sheet stores an explicit 0 (not a
# blank cell) for every station with no parking, so "field is not None" is
# NOT a valid presence check here -- 0 must be treated as absence, same as
# None. Only a positive count or a lot-specific coordinate counts as real
# parking data.
_PARKING_CAPACITY_FIELDS = ("parking_cycle", "parking_motorcycle", "parking_car")


def read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    """Read every data row of the workbook's ``Sheet1`` into a header-keyed dict.

    Blank cells (``None`` or empty string) are coerced to ``None``. Fully
    blank rows (e.g. trailing sheet padding) are skipped.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook["Sheet1"]
        row_iter = worksheet.iter_rows(values_only=True)
        header = [str(cell).strip() for cell in next(row_iter)]
        rows: list[dict[str, Any]] = []
        for values in row_iter:
            if values is None or all(value is None for value in values):
                continue
            record: dict[str, Any] = {}
            for key, value in zip(header, values):
                record[key] = None if value == "" else value
            rows.append(record)
        return rows
    finally:
        workbook.close()


@dataclass
class MatchResult:
    """Outcome of matching xlsx rows against loaded GTFS stops."""

    facilities: list[StationFacility]
    unmatched: list[dict[str, Any]]
    name_matched: int
    coordinate_matched: int


def match_facility_rows(
    xlsx_rows: Sequence[dict[str, Any]], gtfs_stops: Sequence[Stop]
) -> MatchResult:
    """Match xlsx rows to GTFS stops, merging multi-row stations.

    1. Every xlsx row is grouped by normalized ``station_name`` first, so
       stations with multiple parking-lot rows (e.g. Dwarka Mor) merge into
       one :class:`StationFacility` rather than producing duplicates.
    2. Each group is matched by exact normalized-name lookup against the
       GTFS stops; failing that, by nearest GTFS stop within 500 meters of
       the group's own lat/lon.
    3. Groups matching neither way are appended to ``unmatched`` verbatim,
       never dropped silently.
    """
    name_to_stop_id: dict[str, str] = {}
    for stop in gtfs_stops:
        key = normalize_station_name(stop.stop_name)
        if key not in name_to_stop_id:
            name_to_stop_id[key] = stop.stop_id

    groups: dict[str, list[dict[str, Any]]] = {}
    for row in xlsx_rows:
        key = normalize_station_name(str(row.get("station_name") or ""))
        groups.setdefault(key, []).append(row)

    facilities: list[StationFacility] = []
    unmatched: list[dict[str, Any]] = []
    name_matched = 0
    coordinate_matched = 0

    for key, group in groups.items():
        stop_id = name_to_stop_id.get(key)
        match_method: str | None = "name" if stop_id is not None else None

        if stop_id is None:
            nearest_stop_id, nearest_distance = _nearest_stop(group, gtfs_stops)
            if nearest_stop_id is not None and nearest_distance <= _COORDINATE_MATCH_RADIUS_M:
                stop_id = nearest_stop_id
                match_method = "coordinate"

        if stop_id is None or match_method is None:
            unmatched.extend(group)
            continue

        if match_method == "name":
            name_matched += 1
        else:
            coordinate_matched += 1
        facilities.append(_build_facility(stop_id, match_method, group))

    return MatchResult(
        facilities=facilities,
        unmatched=unmatched,
        name_matched=name_matched,
        coordinate_matched=coordinate_matched,
    )


def _nearest_stop(
    rows: Sequence[dict[str, Any]], gtfs_stops: Sequence[Stop]
) -> tuple[str | None, float]:
    """Nearest GTFS stop to any of the group's row coordinates."""
    best_stop_id: str | None = None
    best_distance = math.inf
    for row in rows:
        lat, lon = row.get("lat"), row.get("lon")
        if lat is None or lon is None:
            continue
        for stop in gtfs_stops:
            distance = haversine_meters(float(lat), float(lon), stop.stop_lat, stop.stop_lon)
            if distance < best_distance:
                best_distance = distance
                best_stop_id = stop.stop_id
    return best_stop_id, best_distance


def _build_facility(
    stop_id: str, match_method: str, rows: Sequence[dict[str, Any]]
) -> StationFacility:
    parking_lots = [
        lot for lot in (_build_parking_lot(row) for row in rows) if lot is not None
    ]
    return StationFacility(
        stop_id=stop_id,
        station_code=_first_non_null(rows, "station_code"),
        elevated=_to_bool(_first_non_null(rows, "elevated")),
        toilet=_to_bool(_first_non_null(rows, "toilet")),
        gate_location=_first_non_null(rows, "gate_location"),
        parking_lots=parking_lots or None,
        match_method=match_method,
    )


def _first_non_null(rows: Sequence[dict[str, Any]], key: str) -> Any:
    for row in rows:
        value = row.get(key)
        if value is not None:
            return value
    return None


def _has_parking_data(row: dict[str, Any]) -> bool:
    if not row.get("parking_available"):
        return False
    has_positive_count = any(
        _to_int(row.get(field)) not in (None, 0) for field in _PARKING_CAPACITY_FIELDS
    )
    has_lot_location = row.get("parking_lat") is not None or row.get("parking_lon") is not None
    return has_positive_count or has_lot_location


def _build_parking_lot(row: dict[str, Any]) -> dict[str, Any] | None:
    if not _has_parking_data(row):
        return None
    return {
        "cycle": _to_int(row.get("parking_cycle")),
        "motorcycle": _to_int(row.get("parking_motorcycle")),
        "car": _to_int(row.get("parking_car")),
        "operator": _to_str(row.get("contractor")),
        "contact": _to_str(row.get("contact_number")),
        "lat": _to_float(row.get("parking_lat")),
        "lon": _to_float(row.get("parking_lon")),
    }


def _to_bool(value: Any) -> bool | None:
    return None if value is None else bool(value)


def _to_int(value: Any) -> int | None:
    return None if value is None else int(value)


def _to_float(value: Any) -> float | None:
    return None if value is None else float(value)


def _to_str(value: Any) -> str | None:
    return None if value is None else str(value)


async def load_station_facilities(
    session_factory: SessionFactory, xlsx_path: Path
) -> MatchResult:
    """Match the xlsx dataset against loaded GTFS stops and replace the table.

    Reads every :class:`Stop` row (via :class:`StopRepository`), reads and
    matches the xlsx rows, then wholesale-replaces ``station_facilities``
    inside a single transaction (see the model's docstring). Always returns
    the full :class:`MatchResult` -- including ``unmatched`` -- so the
    caller can report exactly what needs manual attention.
    """
    async with session_factory() as session:
        gtfs_stops = await StopRepository(session).list_all()

    xlsx_rows = read_xlsx_rows(xlsx_path)
    result = match_facility_rows(xlsx_rows, gtfs_stops)

    async with session_factory() as session:
        async with session.begin():
            await StationFacilityRepository(session).replace_all(result.facilities)

    return result
